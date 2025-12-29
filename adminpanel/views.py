from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone

from datetime import datetime
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.contrib.auth.decorators import login_required
from django.utils.text import slugify
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from .utils import (
    get_date_range,
    get_chart_data,
    get_statistics,
    get_best_products,
    get_best_categories
)

from accounts.models import Account
from orders.models import Order,OrderItem
from decimal import Decimal
from orders.utils import get_order_total_discount

# import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


#pdf
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from django.http import HttpResponse



# Create your views here.
@never_cache
def admin_login(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        
        else:
            logout(request) #logout non-admin
            return render(request,'admin_login.html')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        if not email or not password:
            messages.error(request, 'Please fill in all fields.')
            return render(request, 'admin_login.html')     

        user = authenticate(request, email=email, password=password) 

        if user is not None:
            if user.is_superuser:
                login(request,user)
                messages.success(request,'Login successful! Welcome back.')
                return redirect('admin_dashboard')
            else:
                messages.error(request, "You don't have the admin previlages.")
                return render(request,'admin_login.html')
        else:
            messages.error(request,'Invalid email or password.')
            return render(request, 'admin_login.html')   
    return render(request, 'admin_login.html')



@login_required(login_url='admin_login')
@never_cache
def admin_dashboard(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('admin_login')
    
    #get filter from url
    filter_type = request.GET.get('filter', 'week')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # contverts dates if custom filter
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            start_date = None
            end_date = None
            filter_type = 'week'
    
    #get date range based on filter
    date_start, date_end = get_date_range(filter_type, start_date, end_date)

    # get all data range based on filter
    chart_data = get_chart_data(filter_type, date_start, date_end)
    stats = get_statistics(date_start, date_end)
    best_products = get_best_products(date_start, date_end, limit=10)
    best_categories = get_best_categories(date_start, date_end, limit=10)

    
    total_users =  Account.objects.filter(is_superuser=False).count()
    active_users = Account.objects.filter(is_superuser=False, is_active = True).count()
    blocked_users = Account.objects.filter(is_superuser=False, is_active=False).count()


    context = {
        'total_users' : total_users,
        'active_users' : active_users,
        'blocked_users' : blocked_users,

        #chart data
        'chart_labels': chart_data['labels'],
        'chart_sales': chart_data['sales'],
        'chart_orders': chart_data['orders'],

        #statistics
        'total_orders': stats['total_orders'],
        'total_revenue': stats['total_revenue'],
        'total_products_sold': stats['total_products_sold'],
        'average_order_value': stats['average_order_value'],

        #best selling product, categories
        'best_products': best_products,
        'best_categories': best_categories,

        #filter info
        'filter_type': filter_type,
        'start_date': date_start,
        'end_date': date_end,
    }
    return render(request, 'admin_dashboard.html',context)


@login_required(login_url='admin_login')
def user_list(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('admin_login')

    search_query = request.GET.get('search','')

    #get all users excluding superusers
    users = Account.objects.filter(is_superuser=False).order_by('-date_joined')

    if search_query:
        users = users.filter(
            Q(first_name__icontains=search_query)|
            Q(last_name__icontains=search_query)|
            Q(email__icontains=search_query)|
            Q(phone_number__icontains=search_query)
        )        

    paginator = Paginator(users,5) # 10 users per page
    page_number = request.GET.get('page')
    users_page = paginator.get_page(page_number) #paginator method

    context = {
        'users':users_page,
        'search_query' : search_query,
    }
    return render(request,'user_list.html',context)

# Block/Unblock User
@login_required(login_url='admin_login')
def toggle_user_status(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('admin_login')

    user = get_object_or_404(Account,id=user_id, is_superuser=False)

    if user.is_active:
        user.is_active = False
        messages.success(request,f'User {user.email} has been blocked successfully.')
    else:
        user.is_active = True
        messages.success(request,f'User {user.email} has been unblocked successfully.')

    user.save()
    return redirect('user_list')

@login_required(login_url='admin_login')
@require_POST
@never_cache
def admin_logout(request):
    logout(request)
    messages.success(request,"You have been logged out successfuly.")
    return redirect('admin_login')        



VALID_SALES_STATUSES = [
    'confirmed',
    'processing',
    'shipped',
    'out_of_delivery',
    'delivered'
]

def admin_sales_report(request):
    filter_type = request.GET.get('filter', 'week')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    download = request.GET.get('download')

    #convert custom dates
    if start_date and end_date:
        start_date = timezone.datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = timezone.datetime.strptime(end_date, "%Y-%m-%d").date()

    start, end = get_date_range(filter_type, start_date, end_date)

    orders = Order.objects.filter(
        status__in=VALID_SALES_STATUSES,
        created_at__date__range=[start, end]
    )

    total_discount = Decimal("0.00")
    total_sales = Decimal("0.00")

    for order in orders:
        total_discount += get_order_total_discount(order)
        total_sales += order.total_amount

    #aggregations
    summary = {
        "total_orders": orders.count(),
        "total_sales": total_sales,
        "total_discount": total_discount,

    }
    

    if download == "pdf":
        return generate_sales_pdf(orders, summary)
    


    if download == "excel":
        return generate_sales_excel(orders, summary)

    
    for order in orders:
        order.calculated_discount = get_order_total_discount(order)


    context = {
        "orders": orders,
        "summary": summary,
        "filter": filter_type,
        "start_date": start,
        "end_date": end,
    }

    return render(request, "sales_report.html", context)







def generate_sales_pdf(orders, summary, start_date=None, end_date=None):
    """
    Generate a professional sales report PDF with proper table formatting
    """
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    
    # Create PDF with landscape orientation for better table fit
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor("#304B84"),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # ===== HEADER =====
    title = Paragraph("TIMESTAMP SALES REPORT", title_style)
    elements.append(title)
    
    # Date range subtitle
    if start_date and end_date:
        date_range = f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}"
    else:
        date_range = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    
    subtitle = Paragraph(date_range, subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.2*inch))
    
    # ===== SUMMARY CARDS =====
    summary_data = [
        ['TOTAL ORDERS', 'TOTAL SALES', 'TOTAL DISCOUNT', 'NET REVENUE'],
        [
            str(summary.get('total_orders', 0)),
            f"₹{summary.get('total_sales', 0):,.2f}",
            f"₹{summary.get('total_discount', 0):,.2f}",
            f"₹{(summary.get('total_sales', 0) or 0):,.2f}"
        ]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3a5387")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Data row
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1f2937')),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor("#3a5387")),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # ===== ORDERS TABLE =====
    # Table header
    table_data = [
        ['Order ID', 'Customer', 'Date', 'Status', 'Discount', 'Total Amount']
    ]
    
    # Add order rows
    if orders.exists():
        for order in orders:
            discount = get_order_total_discount(order)

            table_data.append([
                order.order_id,
                order.user.email[:25] + '...' if len(order.user.email) > 25 else order.user.email,
                order.created_at.strftime('%d %b %Y'),
                order.get_status_display(),
                f"₹{discount:,.2f}",
                f"₹{order.total_amount:,.2f}"
            ])
    else:
        table_data.append(['', '', 'No orders found', '', '', ''])
    
    # Create table
    orders_table = Table(
        table_data,
        colWidths=[1.5*inch, 2.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.3*inch]
    )
    
    # Table styling
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3a5387')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Data rows styling
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),      # Order ID - left
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),      # Customer - left
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),    # Date - center
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),    # Status - center
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),    # Amounts - right
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#374151')),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Grid and borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#3a5387')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#3a5387')),
    ])
    
    # Alternate row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f9fafb'))
        else:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.white)
    
    orders_table.setStyle(table_style)
    elements.append(orders_table)
    
    #  FOOTER
    elements.append(Spacer(1, 0.3*inch))
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    footer_text = f"Generated on {datetime.now().strftime('%d %B %Y at %I:%M %p')} | Timestamp Store Admin Panel"
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    return response




def generate_sales_excel(orders, summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # ===== TITLE =====
    ws.merge_cells("A1:F1")
    ws["A1"] = "TIMESTAMP SALES REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])

    # ===== SUMMARY =====
    ws.append(["Total Orders", summary["total_orders"]])
    ws.append(["Total Sales", float(summary["total_sales"])])
    ws.append(["Total Discount", float(summary["total_discount"])])
    ws.append([])

    # ===== TABLE HEADER =====
    headers = [
        "Order ID",
        "Customer Email",
        "Status",
        "Discount",
        "Total Amount",
        "Date",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    # ===== TABLE DATA =====
    for order in orders:
        ws.append([
            order.order_id,
            order.user.email,
            order.status,
            float(get_order_total_discount(order)),
            float(order.total_amount),
            order.created_at.strftime("%d-%m-%Y"),
        ])

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)
    return response